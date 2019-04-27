# Excel Source workbook needs to be filled with search results from Facebook
# eg Below results page example would be a search for all people on Facebook who list themselves as currently living in "Newport"
# and who give one of their job titles, in the Facebook work title field, as including the words "staff" and "nurse"
# i.e. most of them will be currently working as a Staff Nurse, maybe a few with previous but not current job as Staff Nurse, however that is rare in practice.
# Note the way of constructing the FB URL search string.
# https://www.facebook.com/search/str/staff+nurse/pages-named/employees/present/str/Newport/pages-named/residents/present/intersect

# Or get FB code for Newport eg Newport, Wales is: 112195725462212
# where we can use: https://www.facebook.com/search/str/staff+nurse/pages-named/employees/present/112195725462212/residents/present/intersect
# for approximately same set of search results, though note there will be some differences as second presumably returns all people whose geo location listed on FB within maximum distance X of centre of Newport, Wales, or something along these lines, whereas the string version will return anyone in world who typed "Newport" as part of location name eg includes Newport in Australia

# To copy and paste results direct from FB, easiest way is:
# 1 Load page formed using URL rules as above
# 2 Keep clicking arrow down or pulling side bar down in web browser until you reach end of results!
# FB results like many sites are lazy loaded with Javascript, so long result list may need keep pressing down for a while!
# 3 Ctrl + A to copy whole page when fully loaded
# IMPORTANT: in Excel sheet, place cursor in Column B row 3 or below and just paste in results
# 4 First several lines will be your own login details FB page generic text etc. - delete these first
# 5 Name this Excel worksheet as something which makes sense for you
# 6 Create a new sheet, and repeat above with the next search string
# 7 Save your file

# NB to run this app from the Windows command prompt, need to activate virtual environment first
# Type facebookenv\Scripts\activate at command prompt with back slashes
# otherwise the virtual environment only loaded packages won't be found and won't run!

import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
# see list of openpyxl Excel style code keywords at: https://openpyxl.readthedocs.io/en/stable/styles.html
from openpyxl.styles import *
# Font, PatternFill, Alignment, Border, Side, Color

class FbResults(object):
    # NB Python syntax reminder: print(<member (obj or function) name>.__doc__) will give the below docstring
    """A class with methods to convert Facebook search results list(s) which have been manually pasted into Excel sheet(s). This will create a neat table of results in a new Excel Workbook. This specific example has been geared to nursing professionals search lists, but could be adjusted for other professions or search fields, in which case these docstring descriptions should be changed :-)"""
    # Cell size constants for results workbook sheet, make big and easier to read
    HEIGHT = 40
    BREAKHEIGHT = 15 # 25/4/19 breaker rows not used at present, thick borders instead
    WIDTH = 30
    # red, cyan, bright green, violet, blue, orange, magenta to colour rows according to job type
    COLOURS = ['FFdc322f', 'FF2aa198', 'FF00FF00', 'FF6c71c4', 'FF268bd2', 'FFcb4b16', 'FFd33682']
  
    def __init__(self, file_in_name, file_out_name, sheet_title):
        self.file_in= file_in_name
        self.file_out = file_out_name
        self.dirpath = os.getcwd()
        self.id_list = []
        self.duplicate_list = []
        self.is_duplicate = False
        self.name_count = 0
        # start at row 4 in result output sheet
        self.count = 4
        self.col_num = 3
        self.job_type = 1
        self.job_numbers = [0, 0, 0, 0, 0, 0, 0]
        self.is_philippino = False
        # Open Workbook for saving results and assign title to active worksheet
        self.result_book = Workbook()
        self.sheet1 = self.result_book.active
        self.sheet1.title = sheet_title

    # returns the completed data list
    def data_to_List(self):
        """Filters relevant Excel source sheet data into Python List, returns this finished List."""
        print("Well, at least the program has started if you're reading this")
        # pythonic: Switch " and ' to avoid escape character \'

        # / or \ ?: Mac and Linux are different, but now Python has pathlib library which can make such problems easier, see this webpage:
        # https://medium.com/@ageitgey/python-3-quick-tip-the-easy-way-to-deal-with-file-paths-on-windows-mac-and-linux-11a072b58d5f

        # load the source Excel book into memory
        fb_sample = load_workbook(self.dirpath + '\\' + self.file_in)
        # absolute example: fb_sample = load_workbook('C:/Users/FB_SampleSource.xlsx') forward slashes!

        # populate Python list with all sheet titles in the manual search results Workbook
        sheet_list = []
        for sheet in fb_sample.worksheets:
            sheet_list.append(sheet.title)


        # Examining Data Format in search results:
        # Most people have Add Friend field then More Options field, then their name i.e. the beginning of record
        # A few have just a blank line before their name
        # jobType sorting criteria:
        # First field after login URL is always the first work title details
        # Some people have another work title field after this, or even a 3rd

        # test print FB data (should be column B) height, sample cell
        for sh in sheet_list:
            print (sh, str(len(fb_sample[sh]['B'])))
            b10 = fb_sample[sh].cell(10,2).value
            print ('Cell B10 this sheet is: ' + str(b10))

        # initialise List to hold all row contents we want
        data_list = []

        # Initialise boolean flag to track where the record for each new person begins in FB data list
        # this will be set false after first line in each profile's data, and true again on last line
        new_person = True

        # Read in all the data pasted into the FB source Workbook into data List
        # Outer loop iterates over each worksheet, inner loop iterates each Column B row in that sheet
        for sheet in sheet_list:
            # find length of Column B
            list_len = len(fb_sample[sheet]['B'])

            # Column B rows
            # When TESTING, can set range(3, x) with x not too large  ********
            # When RUNNING, set range(3, listLen) ***********************
            for x in range(3, list_len):
                this_cell = fb_sample[sheet].cell(row=x, column=2)

                # if new_person flag still true and this cell has a phrase we want record, then this is first cell in new person's record, so add value of True to data list, then set flag false after this item is added to list
                # This way there is no need for 2 Dimensional list eg with profile as D1 and each field as D2
                # (2D array in other language)
                # First check if interested at all in value of this cell:
                # if this_cell.value is not None and this_cell.value != 'More Options' and this_cell.value != 'Add Friend' and this_cell.value != '':

                # None value will raise exception with strip() method, 
                # but pasted lists will contain None (Python's null) value in empty cells
                if this_cell.value is None:
                    stripped_value = ''
                else:
                    stripped_value = this_cell.value.strip()

                if stripped_value not in ('More Options', 'Add Friend', ''):
                    # Mark beginning of new person with a True value in the data list
                    if new_person == True:
                        data_list.append(True)
                    data_list.append(this_cell.value)

                    # if cell contains hyperlink we want, then append this as separate item in the List
                    if this_cell.hyperlink:
                        h = this_cell.hyperlink
                        # don't want any hyperlink for FB's More Options or Lives in values
                        if stripped_value != 'More Options' and stripped_value != 'Lives in':
                            # h.display holds the actual string for hyperlink; split this and get left side profile URL only, no need to splice
                            url_string = h.display
                            if 'ref=br' in url_string:
                                url_string = url_string.split('?ref=br_rs')[0]
                                url_string = url_string.split('&ref=br_rs')[0]
                            data_list.append(url_string)

                    # this cell contains profile data we want, so we are in the person record, so set flag False
                    new_person = False
                else: # this cell data is More Options, Add Friend or nothing, so still not begun new record
                    new_person = True

        # add an extra True value at the end of the list, for the edge case, so last line is formatted properly
        # see below comments expln of this edge case
        data_list.append(True)
        # Now dataList is fully populated with all the profile fields which we want
        return data_list

    @classmethod
    def change_colours(cls, colours = [None, None, None, None, None, None, None]):
        """Change one or more of the row colours across the class, bit naughty as it's supposed to be a constant with capital letters, but just practising using class method"""
        for x in range(0,7):
            if colours[x] != None:
                cls.COLOURS[x] = colours[x]

    def process_next_row(self, name, url):
        # get FB profile string of NEXT person in list
        profile_string = str(url) # type safety convert to string: avoid any None, False, integer type errors
        # cut off 'https://www.facebook.com/' from beginning
        fbook_id = profile_string[25:len(profile_string)]
        if fbook_id in self.id_list:
            # then skip this person because they are a duplicate
            self.is_duplicate = True
            # if this is 3rd or more duplicate, it will already be in the duplicate list, so only want appear once
            if fbook_id not in self.duplicate_list:
                self.duplicate_list.append(name)
                self.duplicate_list.append(fbook_id)
                # Print duplicate names and FB IDs to console
                print("Following person duplicated at least once: ", name, fbook_id)
        else: # this is a new person, not a duplicate
            # increment number of unique people in results
            self.name_count += 1
            self.is_duplicate = False
            # add them to FB IDs List
            self.id_list.append(fbook_id)
            # Reset column to C for the next person's record
            self.col_num = 3
    
            # Now this person's record is printed to the Workbook, reset
            # variables for the next person, who will be displayed in next
            # row of worksheet:
            self.job_type = 1
            self.is_philippino = False
    
            self.count += 1 # move next row

            # base uniform formatting for next row, thick borders to
            # display rows and columns clearly
            borderEdge = Side(border_style = 'thick', color = Color('fdf6e3'))
            self.sheet1.row_dimensions[self.count].height = self.HEIGHT
            for col in range(1, 20):
                cell = self.sheet1.cell(self.count, col)
                cell.fill = PatternFill("solid", fgColor="002b36")
                cell.border = Border(left = borderEdge, right = borderEdge, top = borderEdge, bottom= borderEdge)
                cell.alignment = Alignment(horizontal= 'center', vertical = 'center', wrap_text = True)

    def List_to_workbook(self, data_list):
        """Write the data from the data_list into results Workbook"""
        print("current directory is : " + self.dirpath)
        filepath = self.dirpath + '\\' + self.file_out
        # previous absolute path: filepath = "C:/Users/Sample_FB_Search_Results.xlsx"

        # indicate success in console: save blank results Excel workbook with sheet now named
        self.result_book.save(filepath)
        print('At least results Workbook was saved successfully, empty at this stage though.')

        self.sheet1.column_dimensions['A'].width = 10 # job type number
        self.sheet1.column_dimensions['B'].width = 20 # Philippino or not
        self.sheet1.column_dimensions['D'].width = 45 # login weblink, want it wider

        # Set width of columns C to M in sheet
        # column_dimensions property in openpyxl seems only accept string i.e.
        # column_dimensions['A'] but not [1]
        for col in range(3, 14):
            if col != 4: # Column D set above
                col_string = get_column_letter(col)
                self.sheet1.column_dimensions[col_string].width = self.WIDTH

        # Triple quote comments in Python actually affect memory allocation, so not
        # using here; form good habit
        # Job Type is Integer, no longer simple boolean nurse or not!:
        # 1 Assumed not a real nurse or HCA eg Dental, nursery...
        # 2 Nurse (RGN, staff nurse..), 3 RMN mental nurse,
        # 4 HCA Health Care Assistant, 5 senior/manager nurse, 6 senior HCA,
        # 7 midwife (override other type matches)
        # start with them assumed not a nurse (=1), then check if words in job
        # title qualify them as nurse
        # Put Job Type value in Column A, so can sort easily with sort button in
        # Excel, colour text according to job type

        # isPhilippino - True or False, check for 'Manila' or 'Davao' in cell
        # values, Colour Background different if True, put Philipinno or not Phil in Column B

        # track if this person is already in results from somewhere else
        # is_duplicate = False

        # Initialise List of FB profile IDs, used to detect duplicate profiles
        # appearing in more than one result sheet
        #id_list = []

        #duplicate_list = []
        # first item each new row is the name, second is FB profile URL
        # access URL, chop off the generic front, check if already in idList
        # if not in idList, add it; if in idList, skip this record, add to a duplicate list

        data_len = len(data_list) # total number of items

        # iterate over every field from FB profile data
        for phrase_count in range(0, data_len):
            this_PhRaSe = data_list[phrase_count]
            # Set cell address we want to write to for current phrase
            address = self.sheet1.cell(row=self.count, column=self.col_num)

            # could use while xx: (while condition xx holds)
            if this_PhRaSe != True and not self.is_duplicate:
                this_phrase = this_PhRaSe.lower() # convert string to lower case before checking to make case insensitive
                # Set or reset job type and whether or not Philippino from here:
                if 'studie' in this_phrase and ('davao' in this_phrase or 'manila' in this_phrase):
                    self.is_philippino = True
                # the string ' at ' is in every job field string in a FB results
                # list, with the spaces included
                # NB but using ' at ' with spaces won't work!  I think because
                # Python list comprehension removes spaces from word phrases
                # TODO: remove students to separate jobType ?8?  which overrides
                # all others, by checking if FIRST job listed includes 'student',
                # or 'studies' in a phrase (but not 'studied'!)

                # *********** TODO: HIVE OFF job type checking function
                # check_job_type()
                # Only perform job type checking on phrase if 'at' is in it
                # Set job types hierarchically in this example, so if person has
                # done senior level (higher number) job, won't reset to lower level
                # if they list a 2nd or 3rd job type which is 'lower' level
                job_so_far = self.job_type
                self.job_type = self.check_job_type(job_so_far, this_phrase) # takes 2 args 3 given?self is 3rd given

                address.value = this_PhRaSe # with any upper case letters retained

                # TODO: Currently, job type and is Phillipino values are reset for
                # each phrase in a person's record
                # Ideally, we'd want them set once only at the end of record, which
                # would mean check if next phrase is True, or the edge case, we are
                # at end of data list, check if next to last item in list.
                # In result sheet, fill Column A value with Job Type code
                self.sheet1.cell(row=self.count, column =1).value = self.job_type
                # Column B is Phillippino or not
                if self.is_philippino:
                    self.sheet1.cell(self.count, 2).value = "Philippino"
                else:
                    self.sheet1.cell(self.count, 2).value = "Not Philippino"

                # if this field in data list is a URL, then add the hyperlink to
                # this URL in the Excel sheet
                # so can click straight to FB profile from the sheet
                if '://' in this_PhRaSe:
                    address.hyperlink = this_PhRaSe
        
               # move next column right for next phrase in the list
                self.col_num += 1

            # If this is beginning of new person record, process accordingly.
            # Will skip above to arrive here, if capitalised phrase True OR is_duplicate is True OR both...
            elif this_PhRaSe == True: # last record should be True for the edge case
        
                # current record definitely finalised, set row text colour according to job type
                # list for job types 1 - 7, so List index 0 - 6
                for col in range(1, 20):
                    self.sheet1.cell(self.count, col).font = Font(name='Verdana', bold = True, color = self.COLOURS[self.job_type - 1])
                # add this person to running count for their job type
                self.job_numbers[self.job_type - 1] += 1
        
                # Next row is beginning of new person record, prepare for NEXT row now
                # Except Edge Case: Last item in phrase list is dummy True value, with no record after it
                try:
                    next_name = data_list[phrase_count + 1]
                    next_url = data_list[phrase_count + 2]
                    self.process_next_row(next_name, next_url)
                    # call big hived off function here, NB this won't be called if above line index errors anyway
                    # call with next_name, next_url to avoid passing whole datalist
                except IndexError:
                    print('Reached end of the list here at number ' + str(phrase_count) + ' phrase in the list, out of range error reached, this should happen at end of list anyway. That number should be total number of Excel rows in source sheets we actually want copied to the final table, if this is working correctly.')
    
    # TODO: start with just specific strings filter here, consider fuzzy match with fuzzywuzzy package later
    @staticmethod
    def check_job_type(job_type, this_phrase):
        """Checks if main job type should change, based on current phrase, returns job type."""
        if (job_type < 7) and 'at' in this_phrase and not any(_ in this_phrase for _ in ['studie', 'nursery', 'dental', 'nursary', 'vet']): # once set as Midwife cannot reset
            if 'midwife' in this_phrase or 'natal' in this_phrase:
                job_type = 7
    
            if job_type < 6: # Senior nurse cannot reset, but if 5 can reset to 6
                if 'senior' in this_phrase or 'manager' in this_phrase:
                    if any(a in this_phrase for a in ['hca', 'hcsw', 'health', 'care assistant', 'h.c.a.', 'support worker']):
                        job_type = 5
                    else:
                        job_type = 6
    
                if job_type < 5:
                    if 'assistant' not in this_phrase and any(b in this_phrase for b in ['rmn', 'r.m.n.', 'mental', 'psychiatric']): # captures 'mental health nurse'
                        job_type = 4
                    if job_type < 4: # RMN or above can't change
                        if any(c in this_phrase for c in ['nurse', 'rgn', 'r.g.n.', 'nursing']) and 'assistant' not in this_phrase: # captures 'works at nurse', 'nursing at'
                            job_type = 3
                        if job_type == 1: # was if < 3, but if it's 2 already, no need check again here
                            if 'works at' in this_phrase and any(d in this_phrase for d in ['nhs', 'n.h.s.', 'n.h.s', 'hospital']):
                                job_type = 2
                            if any(e in this_phrase for e in ['support worker', 'hcsw', 'h.c.s.w.', 'hca', 'h.c.a.', 'h.c.a', 'assistant', 'health', 'nhs', 'n.h.s.', 'n.h.s']): # this should capture 'health care assistant'
                                job_type = 2
        return job_type

    def summarise_sheet(self):
        """Write summary totals etc. top of sheet"""
        self.sheet1.cell(1, 3).value = 'There are a total of ' + str(self.name_count) + ' names with unique Facebook logins in this list.'
        # display job type totals at top of Excel sheet
        job_names = ['Not Nurse', 'HCA', 'Nurse', 'Mental', 'Senior HCA/Non-nurse', 'Senior Nurse', 'Midwife']
        for z in range (1, 7):
            self.sheet1.cell(1, z + 5).value = job_names[z - 1]
            self.sheet1.cell(2, z + 5).value = self.job_numbers[z - 1]
    
        # display number of duplicate profiles removed from total original results set
        # NB each duplicate has name and ID both added to list, so total number duplicates is dup_len/2
        dup_len = len(self.duplicate_list)
        s = self.sheet1
        s.merge_cells('C2:E3')
        s.cell(2,3).alignment = Alignment(wrap_text= True)
        s.row_dimensions[2].height = 25
        s.cell(2, 3).value = 'There were a total of ' + str(int(dup_len/2)) + ' duplicate Facebook logins removed from the original list, NB any duplicated login appeared 2 or more times in the original merged list. See second sheet for the names and profile IDs of these duplicates.'
        # Create a new sheet Duplicates
        self.result_book.create_sheet(title = "Duplicates")
        dupSheet = self.result_book['Duplicates']
        dupSheet.column_dimensions['A'].width = 30
        dupSheet.column_dimensions['B'].width = 30
        # Populate Duplicates sheet from the list
        for ind in range(0, dup_len, 2):
            dupSheet.cell(ind/2 + 1, 1).value = self.duplicate_list[ind] # name in Column A
            dupSheet.cell(ind/2 + 1, 2).value = self.duplicate_list[ind + 1] # ID in Column B
    
    def save(self, new_outfile = None):
        """Option to save as new output file name"""
        if new_outfile is None:
            self.result_book.save(self.dirpath + '\\' + self.file_out)
        else:
            self.result_book.save(self.dirpath + '\\' + new_outfile)

nurses = FbResults('\FB_SampleSource.xlsx', '\FB_Processed_Search_Results.xlsx', 'NursesAndHCAs')
data_list = nurses.data_to_List()
nurses.List_to_workbook(data_list)
nurses.summarise_sheet()
# Remember to save the workbook or it's all been for nothing!
nurses.save()

# As this is console app, extra indication in command prompt that program finished executing
print ('Finished now, check the file')
